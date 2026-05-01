#!/usr/bin/env python3
"""Generate a WooCommerce-ready product import CSV from the NeoGen catalog."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_INPUT = Path("data/catalogs/master/Neogen_Master_Catalog_Blueprint.xlsx")
DEFAULT_OUTPUT = Path("output/spreadsheet/woocommerce_ready_import.csv")
DEFAULT_BUY_LINKS_OUTPUT = Path("output/spreadsheet/woocommerce_buy_links.csv")
DEFAULT_SHEET = "1. Master Catalog"
DEFAULT_SITE_URL = "https://neogen.store"

REQUIRED_COLUMNS = {
    "ID",
    "SKU",
    "Product Name (EN)",
    "Category",
    "Condition",
    "Shipping Class",
    "Regular Price (SAR)",
    "Sale Price (SAR)",
    "Warranty (months)",
    "Warranty Type",
}

OPTIONAL_COPY_COLUMNS = {
    "GTIN, UPC, EAN, or ISBN": "GTIN / EAN / UPC",
    "Brands": "Brand",
    "Weight (kg)": "Weight (kg)",
    "Length (cm)": "Length (cm)",
    "Width (cm)": "Width (cm)",
    "Height (cm)": "Height (cm)",
    "Shipping class": "Shipping Class",
    "Meta: _ng_ar_title": "Product Name (AR)",
    "Meta: _ng_ar_description": "Short Description (AR)",
    "Meta: _ng_model_number": "Model Number",
    "Meta: _ng_country_of_origin": "Country of Origin",
    "Meta: _ng_hs_code": "HS Code",
    "Meta: _ng_voltage": "Voltage",
    "Meta: _ng_lifecycle_status": "Lifecycle Status",
}

OUTPUT_COLUMNS = [
    "ID",
    "Type",
    "SKU",
    "GTIN, UPC, EAN, or ISBN",
    "Name",
    "Published",
    "Is featured?",
    "Visibility in catalog",
    "Sale price",
    "Regular price",
    "Categories",
    "Brands",
    "Shipping class",
    "Weight (kg)",
    "Length (cm)",
    "Width (cm)",
    "Height (cm)",
    "Meta: _ng_warranty",
    "Meta: _ng_tech_specs",
    "Meta: _ng_ar_title",
    "Meta: _ng_ar_description",
    "Meta: _ng_model_number",
    "Meta: _ng_country_of_origin",
    "Meta: _ng_hs_code",
    "Meta: _ng_voltage",
    "Meta: _ng_lifecycle_status",
]

BUY_LINK_COLUMNS = [
    "ID",
    "SKU",
    "Name",
    "Regular price",
    "Sale price",
    "Categories",
    "Product ID URL",
    "Add to cart URL",
    "Cart buy URL",
    "Checkout buy URL",
]


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def clean(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).split())


def money(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, (int, float)):
        return f"{value:g}"
    return clean(value)


def make_warranty(row: dict[str, Any]) -> str:
    months = row.get("Warranty (months)")
    warranty_type = row.get("Warranty Type")

    if not is_blank(months):
        try:
            month_text = str(int(float(months)))
        except (TypeError, ValueError):
            month_text = clean(months)

        if not is_blank(warranty_type):
            return f"{month_text} Months {clean(warranty_type)}"
        return f"{month_text} Months Standard Warranty"

    return "Standard NGS Warranty"


def make_tech_specs(row: dict[str, Any]) -> str:
    specs: list[str] = []
    for label, column in (
        ("Brand", "Brand"),
        ("Model", "Model Number"),
        ("Condition", "Condition"),
        ("Class", "Shipping Class"),
        ("Voltage", "Voltage"),
    ):
        value = row.get(column)
        if not is_blank(value):
            specs.append(f"{label}: {clean(value)}")
    return " | ".join(specs) if specs else "Standard Specifications"


def load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean(header) for header in header_row]
    missing = sorted(REQUIRED_COLUMNS - set(headers))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(not is_blank(value) for value in row.values()):
            rows.append(row)
    return rows


def make_woo_row(source: dict[str, Any]) -> dict[str, str | int]:
    target: dict[str, str | int] = {column: "" for column in OUTPUT_COLUMNS}

    target.update(
        {
            "ID": clean(source.get("ID")),
            "Type": "simple",
            "SKU": clean(source.get("SKU")),
            "Name": clean(source.get("Product Name (EN)")),
            "Published": 1,
            "Is featured?": 0,
            "Visibility in catalog": "visible",
            "Sale price": money(source.get("Sale Price (SAR)")),
            "Regular price": money(source.get("Regular Price (SAR)")),
            "Categories": clean(source.get("Category")),
            "Meta: _ng_warranty": make_warranty(source),
            "Meta: _ng_tech_specs": make_tech_specs(source),
        }
    )

    for output_column, input_column in OPTIONAL_COPY_COLUMNS.items():
        target[output_column] = clean(source.get(input_column))

    return target


def normalize_site_url(site_url: str) -> str:
    return site_url.strip().rstrip("/")


def make_buy_link_row(woo_row: dict[str, str | int], site_url: str) -> dict[str, str | int]:
    product_id = clean(woo_row["ID"])
    return {
        "ID": product_id,
        "SKU": woo_row["SKU"],
        "Name": woo_row["Name"],
        "Regular price": woo_row["Regular price"],
        "Sale price": woo_row["Sale price"],
        "Categories": woo_row["Categories"],
        "Product ID URL": f"{site_url}/?p={product_id}",
        "Add to cart URL": f"{site_url}/?add-to-cart={product_id}",
        "Cart buy URL": f"{site_url}/cart/?add-to-cart={product_id}&quantity=1",
        "Checkout buy URL": f"{site_url}/checkout/?add-to-cart={product_id}&quantity=1",
    }


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str | int]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def generate_csv(
    input_path: Path,
    output_path: Path,
    sheet_name: str,
    buy_links_output: Path | None = None,
    site_url: str = DEFAULT_SITE_URL,
) -> tuple[int, int]:
    rows = load_rows(input_path, sheet_name)
    woo_rows = [make_woo_row(row) for row in rows]

    write_csv(output_path, OUTPUT_COLUMNS, woo_rows)

    skus = [row["SKU"] for row in woo_rows if row["SKU"]]
    duplicate_skus = len(skus) - len(set(skus))
    if duplicate_skus:
        raise ValueError(f"Generated CSV has {duplicate_skus} duplicate SKU values")

    buy_link_count = 0
    if buy_links_output is not None:
        normalized_site_url = normalize_site_url(site_url)
        buy_link_rows = [make_buy_link_row(row, normalized_site_url) for row in woo_rows]
        write_csv(buy_links_output, BUY_LINK_COLUMNS, buy_link_rows)
        buy_link_count = len(buy_link_rows)

    return len(woo_rows), buy_link_count


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a WooCommerce import CSV from Neogen_Master_Catalog_Blueprint.xlsx."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--buy-links-output", type=Path, default=DEFAULT_BUY_LINKS_OUTPUT)
    parser.add_argument("--site-url", default=DEFAULT_SITE_URL)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--no-buy-links", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    buy_links_output = None if args.no_buy_links else args.buy_links_output
    count, buy_link_count = generate_csv(
        args.input,
        args.output,
        args.sheet,
        buy_links_output=buy_links_output,
        site_url=args.site_url,
    )
    print(f"Generated {args.output} with {count} products.")
    if buy_links_output is not None:
        print(f"Generated {buy_links_output} with {buy_link_count} buy links.")


if __name__ == "__main__":
    main()
