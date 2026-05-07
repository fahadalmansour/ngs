"""
Apply scraped Amazon SA prices into the master xlsx + supplier prices into the supplier matrix.

Reads:
  - output/sync/price-sync-live-inscope-2026-05-07.csv  (scraper output)
  - data/catalogs/master/Neogen_Master_Catalog_Blueprint.xlsx (master catalog)

Writes (in place, with backup):
  - data/catalogs/master/Neogen_Master_Catalog_Blueprint.xlsx (Amazon SA Ref Price column updated for in-scope SKUs)

Then user runs:
  npm run sourcing:generate
  npm run price:guard
to refresh downstream artifacts.

Per the truth-model rule, this script ONLY writes confirmed values. Empty results in
the scraper output are left alone (existing master cell preserved).
"""

import csv, sys, shutil, datetime, json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCRAPE = ROOT / "output/sync/price-sync-live-inscope-2026-05-07.csv"
MASTER = ROOT / "data/catalogs/master/Neogen_Master_Catalog_Blueprint.xlsx"

if not SCRAPE.exists():
    sys.exit(f"Missing scraper output: {SCRAPE}")
if not MASTER.exists():
    sys.exit(f"Missing master xlsx: {MASTER}")

# Backup
ts = datetime.datetime.utcnow().strftime("%Y%m%d-%H%M%S")
backup = MASTER.with_suffix(MASTER.suffix + f".backup-{ts}")
shutil.copy2(MASTER, backup)
print(f"Backed up master → {backup.name}")

# Read scraper output
with SCRAPE.open(encoding="utf-8") as f:
    scraped = list(csv.DictReader(f))
print(f"Scraper rows: {len(scraped)}")

# Open master xlsx
wb = load_workbook(MASTER)
ws = wb["1. Master Catalog"]
header = [c.value for c in ws[1]]

sku_col = header.index("SKU") + 1
amzn_col = header.index("Amazon SA Ref Price") + 1
amzn_url_col = header.index("Amazon Ref URL") + 1 if "Amazon Ref URL" in header else None
amzn_date_col = header.index("Amazon Ref Date") + 1 if "Amazon Ref Date" in header else None

# Build SKU → row map
sku_to_row = {}
for r in range(2, ws.max_row + 1):
    sku = ws.cell(row=r, column=sku_col).value
    if sku:
        sku_to_row[str(sku).strip()] = r

today_iso = datetime.date.today().isoformat()

# Sanity-check thresholds — the scraper's `.a-price-whole` selector sometimes
# grabs sponsored ads or filter widgets instead of the real product card.
# A scraped price of 65 SAR for a 5000-SAR firewall is almost certainly wrong.
SUSPECT_LOW_RATIO = 0.05    # < 5% of NeoGen list = reject (wrong listing)
REVIEW_LOW_RATIO  = 0.20    # 5-20% = flag for manual review (don't auto-apply)
REVIEW_HIGH_RATIO = 5.00    # > 5x = also suspect (matched a way-pricier item)

# Read the master Regular Price for ratio comparison
reg_price_col = header.index("Regular Price (SAR)") + 1
reg_prices = {}
for r in range(2, ws.max_row + 1):
    sku = ws.cell(row=r, column=sku_col).value
    rp = ws.cell(row=r, column=reg_price_col).value
    if sku and rp:
        try:
            reg_prices[str(sku).strip()] = float(rp)
        except (TypeError, ValueError):
            pass

applied = 0
skipped_no_value = 0
skipped_no_match = 0
skipped_already_set = 0
rejected_too_low = 0
flagged_for_review = 0
applied_log = []
review_log = []

for s in scraped:
    sku = (s.get("SKU") or "").strip()
    amzn = (s.get("Amazon_SA_Ref_SAR") or "").strip()
    if not sku:
        continue
    if not amzn:
        skipped_no_value += 1
        continue
    if sku not in sku_to_row:
        skipped_no_match += 1
        continue
    r = sku_to_row[sku]
    existing = ws.cell(row=r, column=amzn_col).value
    if existing:
        skipped_already_set += 1
        continue
    try:
        amzn_val = float(amzn)
    except ValueError:
        skipped_no_value += 1
        continue

    # Sanity check vs NeoGen Regular Price
    nk_list = reg_prices.get(sku)
    ratio = (amzn_val / nk_list) if nk_list else None
    if ratio is not None and ratio < SUSPECT_LOW_RATIO:
        rejected_too_low += 1
        review_log.append({"sku": sku, "amazon_sa_ref": amzn_val, "neogen_list": nk_list,
                           "ratio": round(ratio, 3), "decision": "rejected_too_low"})
        continue
    if ratio is not None and (ratio < REVIEW_LOW_RATIO or ratio > REVIEW_HIGH_RATIO):
        flagged_for_review += 1
        review_log.append({"sku": sku, "amazon_sa_ref": amzn_val, "neogen_list": nk_list,
                           "ratio": round(ratio, 3), "decision": "flagged_for_manual_review"})
        continue

    # Apply
    ws.cell(row=r, column=amzn_col).value = amzn_val
    if amzn_date_col:
        ws.cell(row=r, column=amzn_date_col).value = today_iso
    if amzn_url_col:
        ws.cell(row=r, column=amzn_url_col).value = "scraped via scripts/live-market-sync-inscope.js (2026-05-07)"
    applied += 1
    applied_log.append({"sku": sku, "amazon_sa_ref": amzn_val,
                        "neogen_list": nk_list, "ratio": round(ratio, 3) if ratio else None})

wb.save(MASTER)

print(f"\nApplied:                                    {applied}")
print(f"Rejected (scraped < 5% of NeoGen list):      {rejected_too_low}")
print(f"Flagged for review (5-20% or >500%):         {flagged_for_review}")
print(f"Skipped (no Amazon SA value scraped):        {skipped_no_value}")
print(f"Skipped (already had a price in master):     {skipped_already_set}")
print(f"Skipped (SKU not in master):                 {skipped_no_match}")

audit_log = ROOT / "docs/feasibility" / f"2026-05-07-market-sync-applied.json"
audit_log.write_text(json.dumps({
    "applied_at": datetime.datetime.utcnow().isoformat() + "Z",
    "source_csv": str(SCRAPE.relative_to(ROOT)),
    "master_xlsx": str(MASTER.relative_to(ROOT)),
    "backup": backup.name,
    "applied_count": applied,
    "rejected_too_low": rejected_too_low,
    "flagged_for_review": flagged_for_review,
    "skipped_no_value": skipped_no_value,
    "skipped_already_set": skipped_already_set,
    "skipped_no_match": skipped_no_match,
    "thresholds": {
        "suspect_low_ratio": SUSPECT_LOW_RATIO,
        "review_low_ratio":  REVIEW_LOW_RATIO,
        "review_high_ratio": REVIEW_HIGH_RATIO,
    },
    "applied_skus": applied_log,
    "review_skus":  review_log,
}, indent=2))
print(f"Audit log → {audit_log.relative_to(ROOT)}")
print()
print("Next: npm run woo:generate && npm run sourcing:generate && npm run price:guard")
