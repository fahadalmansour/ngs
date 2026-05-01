#!/usr/bin/env python3
"""Generate supplier sourcing sheets with five supplier targets per product."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus

from openpyxl import load_workbook


DEFAULT_INPUT = Path("data/catalogs/master/Neogen_Master_Catalog_Blueprint.xlsx")
DEFAULT_LONG_OUTPUT = Path("output/spreadsheet/supplier_sourcing_matrix.csv")
DEFAULT_WIDE_OUTPUT = Path("output/spreadsheet/supplier_sourcing_wide.csv")
DEFAULT_SHEET = "1. Master Catalog"

REQUIRED_COLUMNS = {
    "ID",
    "SKU",
    "Product Name (EN)",
    "Category",
    "Regular Price (SAR)",
    "Sale Price (SAR)",
}

SUPPLIERS = [
    {
        "name": "AliExpress",
        "type": "Marketplace supplier search",
        "currency": "USD",
        "search": "https://www.aliexpress.com/wholesale?SearchText={query}",
        "notes": "Match exact product manually, then fill Buy URL and Supplier Price.",
    },
    {
        "name": "Alibaba",
        "type": "B2B supplier search",
        "currency": "USD",
        "search": "https://www.alibaba.com/trade/search?SearchText={query}",
        "notes": "Check MOQ, shipping terms, and verified supplier status before buying.",
    },
    {
        "name": "Amazon.sa",
        "type": "Local market reference / supplier option",
        "currency": "SAR",
        "search": "https://www.amazon.sa/s?k={query}",
        "notes": "Uses existing Amazon reference URL and price when the workbook has them.",
    },
    {
        "name": "Noon",
        "type": "Local market reference / supplier option",
        "currency": "SAR",
        "search": "https://www.noon.com/saudi-en/search/?q={query}",
        "notes": "Verify seller rating, warranty, and marketplace fulfillment before buying.",
    },
    {
        "name": "B&H Photo",
        "type": "US retailer / distributor search",
        "currency": "USD",
        "search": "https://www.bhphotovideo.com/c/search?Ntt={query}&N=0&InitialSearch=yes&sts=ma",
        "notes": "Best for cameras, networking, creators, and computer hardware.",
    },
]

LONG_COLUMNS = [
    "ID",
    "SKU",
    "Product Name",
    "Category",
    "Regular Price SAR",
    "Sale Price SAR",
    "Supplier Rank",
    "Supplier Name",
    "Supplier Type",
    "Supplier Search URL",
    "Supplier Buy URL",
    "Supplier Price",
    "Currency",
    "Price Source",
    "Notes",
]

WIDE_BASE_COLUMNS = [
    "ID",
    "SKU",
    "Product Name",
    "Category",
    "Regular Price SAR",
    "Sale Price SAR",
]


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def clean(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).split())


def money(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, (int, float)):
        return f"{value:g}"
    return clean(value)


def load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    sheet = workbook[sheet_name]
    headers = [clean(header) for header in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    missing = sorted(REQUIRED_COLUMNS - set(headers))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(not is_blank(value) for value in row.values()):
            rows.append(row)
    return rows


def query_for(row: dict[str, Any]) -> str:
    parts = [
        clean(row.get("Brand")),
        clean(row.get("Model Number")),
        clean(row.get("MPN")),
        clean(row.get("Product Name (EN)")),
    ]
    query = " ".join(part for part in parts if part)
    return quote_plus(query or clean(row.get("SKU")))


def base_product_row(row: dict[str, Any]) -> dict[str, str]:
    return {
        "ID": clean(row.get("ID")),
        "SKU": clean(row.get("SKU")),
        "Product Name": clean(row.get("Product Name (EN)")),
        "Category": clean(row.get("Category")),
        "Regular Price SAR": money(row.get("Regular Price (SAR)")),
        "Sale Price SAR": money(row.get("Sale Price (SAR)")),
    }


def make_supplier_row(row: dict[str, Any], supplier: dict[str, str], rank: int) -> dict[str, str | int]:
    query = query_for(row)
    search_url = supplier["search"].format(query=query)
    buy_url = ""
    supplier_price = ""
    price_source = "Needs supplier verification"

    if supplier["name"] == "Amazon.sa":
        buy_url = clean(row.get("Amazon Ref URL"))
        supplier_price = money(row.get("Amazon SA Ref Price"))
        if buy_url or supplier_price:
            price_source = "Existing workbook Amazon reference"

    return {
        **base_product_row(row),
        "Supplier Rank": rank,
        "Supplier Name": supplier["name"],
        "Supplier Type": supplier["type"],
        "Supplier Search URL": search_url,
        "Supplier Buy URL": buy_url,
        "Supplier Price": supplier_price,
        "Currency": supplier["currency"],
        "Price Source": price_source,
        "Notes": supplier["notes"],
    }


def make_long_rows(rows: list[dict[str, Any]]) -> list[dict[str, str | int]]:
    supplier_rows: list[dict[str, str | int]] = []
    for row in rows:
        for rank, supplier in enumerate(SUPPLIERS, start=1):
            supplier_rows.append(make_supplier_row(row, supplier, rank))
    return supplier_rows


def make_wide_rows(rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, str]]]:
    supplier_columns: list[str] = []
    for rank in range(1, len(SUPPLIERS) + 1):
        supplier_columns.extend(
            [
                f"Supplier {rank} Name",
                f"Supplier {rank} Search URL",
                f"Supplier {rank} Buy URL",
                f"Supplier {rank} Price",
                f"Supplier {rank} Currency",
                f"Supplier {rank} Price Source",
            ]
        )

    wide_rows: list[dict[str, str]] = []
    for row in rows:
        wide_row = base_product_row(row)
        for rank, supplier in enumerate(SUPPLIERS, start=1):
            supplier_row = make_supplier_row(row, supplier, rank)
            wide_row.update(
                {
                    f"Supplier {rank} Name": clean(supplier_row["Supplier Name"]),
                    f"Supplier {rank} Search URL": clean(supplier_row["Supplier Search URL"]),
                    f"Supplier {rank} Buy URL": clean(supplier_row["Supplier Buy URL"]),
                    f"Supplier {rank} Price": clean(supplier_row["Supplier Price"]),
                    f"Supplier {rank} Currency": clean(supplier_row["Currency"]),
                    f"Supplier {rank} Price Source": clean(supplier_row["Price Source"]),
                }
            )
        wide_rows.append(wide_row)

    return WIDE_BASE_COLUMNS + supplier_columns, wide_rows


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def generate_outputs(input_path: Path, sheet_name: str, long_output: Path, wide_output: Path) -> tuple[int, int]:
    product_rows = load_rows(input_path, sheet_name)
    long_rows = make_long_rows(product_rows)
    wide_columns, wide_rows = make_wide_rows(product_rows)

    write_csv(long_output, LONG_COLUMNS, long_rows)
    write_csv(wide_output, wide_columns, wide_rows)

    return len(product_rows), len(long_rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate five supplier sourcing targets per NeoGen product."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--long-output", type=Path, default=DEFAULT_LONG_OUTPUT)
    parser.add_argument("--wide-output", type=Path, default=DEFAULT_WIDE_OUTPUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    product_count, supplier_count = generate_outputs(
        args.input,
        args.sheet,
        args.long_output,
        args.wide_output,
    )
    print(f"Generated {args.long_output} with {supplier_count} supplier rows.")
    print(f"Generated {args.wide_output} with {product_count} product rows.")


if __name__ == "__main__":
    main()
