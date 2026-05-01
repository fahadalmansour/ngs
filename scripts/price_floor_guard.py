#!/usr/bin/env python3
"""Audit WooCommerce prices against available real-price floors."""

from __future__ import annotations

import argparse
import csv
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_MASTER = Path("data/catalogs/master/Neogen_Master_Catalog_Blueprint.xlsx")
DEFAULT_WOO_INPUT = Path("output/spreadsheet/woocommerce_ready_import.csv")
DEFAULT_CHECKED_OUTPUT = Path("output/spreadsheet/woocommerce_ready_import_price_floor_checked.csv")
DEFAULT_STRICT_OUTPUT = Path("output/spreadsheet/woocommerce_ready_import_safe_verified_only.csv")
DEFAULT_AUDIT_OUTPUT = Path("output/spreadsheet/price_floor_audit.csv")
DEFAULT_EXCEPTIONS_OUTPUT = Path("output/spreadsheet/price_floor_exceptions.csv")
DEFAULT_SUPPLIER_MATRIX = Path("output/spreadsheet/supplier_sourcing_matrix.csv")
DEFAULT_SHEET = "1. Master Catalog"
DEFAULT_USD_TO_SAR = 3.7502

MONEY_RE = re.compile(r"\$\s*([0-9]+(?:\.[0-9]+)?)")
GIFT_CATEGORY = "Gift Cards & Software Keys"

AUDIT_COLUMNS = [
    "ID",
    "SKU",
    "Name",
    "Categories",
    "Regular price",
    "Sale price",
    "Effective sell price SAR",
    "Real price floor SAR",
    "Real price source",
    "Reference URL",
    "Status",
    "Risk",
    "Adjustment",
]


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def clean(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).split())


def parse_money(value: Any) -> float | None:
    if is_blank(value):
        return None
    try:
        return float(str(value).strip().replace(",", ""))
    except ValueError:
        return None


def money(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}".rstrip("0").rstrip(".")


def load_master_refs(path: Path, sheet_name: str) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    headers = [clean(header) for header in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: dict[str, dict[str, str]] = {}

    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        sku = clean(row.get("SKU"))
        if not sku:
            continue
        rows[sku] = {
            "amazon_price": clean(row.get("Amazon SA Ref Price")),
            "amazon_url": clean(row.get("Amazon Ref URL")),
        }
    return rows


def load_supplier_refs(path: Path, usd_to_sar: float) -> dict[str, list[dict[str, str | float]]]:
    if not path.exists():
        return {}

    refs: dict[str, list[dict[str, str | float]]] = {}
    with path.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            price = parse_money(row.get("Supplier Price"))
            if price is None:
                continue

            currency = clean(row.get("Currency")).upper()
            if currency == "USD":
                price_sar = price * usd_to_sar
            elif currency == "SAR":
                price_sar = price
            else:
                continue

            sku = clean(row.get("SKU"))
            if not sku:
                continue

            refs.setdefault(sku, []).append(
                {
                    "price_sar": price_sar,
                    "source": f"{clean(row.get('Supplier Name'))} supplier matrix",
                    "url": clean(row.get("Supplier Buy URL") or row.get("Supplier Search URL")),
                }
            )
    return refs


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def rows_match(left: list[dict[str, str]], right: list[dict[str, str]], columns: list[str]) -> bool:
    if len(left) != len(right):
        return False
    for left_row, right_row in zip(left, right):
        for column in columns:
            if clean(left_row.get(column)) != clean(right_row.get(column)):
                return False
    return True


def effective_price(row: dict[str, str]) -> float | None:
    sale = parse_money(row.get("Sale price"))
    if sale is not None:
        return sale
    return parse_money(row.get("Regular price"))


def real_price_refs(
    row: dict[str, str],
    master_refs: dict[str, dict[str, str]],
    supplier_refs: dict[str, list[dict[str, str | float]]],
    usd_to_sar: float,
) -> list[dict[str, str | float]]:
    refs: list[dict[str, str | float]] = []
    name = clean(row.get("Name"))
    face_match = MONEY_RE.search(name)
    if face_match:
        face_usd = float(face_match.group(1))
        refs.append(
            {
                "price_sar": face_usd * usd_to_sar,
                "source": f"USD face value from product name (${face_usd:g})",
                "url": "",
            }
        )

    sku = clean(row.get("SKU"))
    master = master_refs.get(sku, {})
    amazon_price = parse_money(master.get("amazon_price"))
    if amazon_price is not None:
        refs.append(
            {
                "price_sar": amazon_price,
                "source": "Amazon.sa reference from master catalog",
                "url": master.get("amazon_url", ""),
            }
        )

    refs.extend(supplier_refs.get(sku, []))
    return refs


def classify_missing_reference(row: dict[str, str], price: float | None) -> tuple[str, str]:
    category = clean(row.get("Categories"))
    if GIFT_CATEGORY in category:
        return "NEEDS_REFERENCE", "High"
    if price is None or price <= 0:
        return "NEEDS_REFERENCE", "High"
    return "NO_REAL_PRICE_REFERENCE", "Medium"


def apply_floor(row: dict[str, str], floor: float) -> tuple[dict[str, str], str]:
    adjusted = dict(row)
    floor_price = math.ceil(floor)
    sale = parse_money(adjusted.get("Sale price"))
    regular = parse_money(adjusted.get("Regular price"))

    if sale is not None:
        adjusted["Sale price"] = str(floor_price)
        if regular is None or regular < floor_price:
            adjusted["Regular price"] = str(floor_price)
        return adjusted, f"Raised sale price to {floor_price} SAR"

    adjusted["Regular price"] = str(floor_price)
    return adjusted, f"Raised regular price to {floor_price} SAR"


def audit_rows(
    woo_rows: list[dict[str, str]],
    master_refs: dict[str, dict[str, str]],
    supplier_refs: dict[str, list[dict[str, str | float]]],
    usd_to_sar: float,
    unpublish_missing_reference: bool = False,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    checked_rows: list[dict[str, str]] = []
    strict_rows: list[dict[str, str]] = []
    audit: list[dict[str, str]] = []

    for row in woo_rows:
        refs = real_price_refs(row, master_refs, supplier_refs, usd_to_sar)
        price = effective_price(row)
        checked = dict(row)
        strict = dict(row)
        floor = None
        source = ""
        url = ""
        status = "PASS"
        risk = "Low"
        adjustment = ""

        if refs:
            best_ref = max(refs, key=lambda ref: float(ref["price_sar"]))
            floor = float(best_ref["price_sar"])
            source = clean(best_ref["source"])
            url = clean(best_ref.get("url", ""))
            if price is None or price < floor:
                checked, adjustment = apply_floor(row, floor)
                strict = dict(checked)
                status = "FIXED"
                risk = "High"
        else:
            status, risk = classify_missing_reference(row, price)
            if unpublish_missing_reference:
                strict["Published"] = "0"
                adjustment = "Strict output unpublished until real price reference is filled"

        checked_rows.append(checked)
        strict_rows.append(strict)
        audit.append(
            {
                "ID": clean(row.get("ID")),
                "SKU": clean(row.get("SKU")),
                "Name": clean(row.get("Name")),
                "Categories": clean(row.get("Categories")),
                "Regular price": clean(row.get("Regular price")),
                "Sale price": clean(row.get("Sale price")),
                "Effective sell price SAR": money(price),
                "Real price floor SAR": money(floor),
                "Real price source": source,
                "Reference URL": url,
                "Status": status,
                "Risk": risk,
                "Adjustment": adjustment,
            }
        )

    return checked_rows, strict_rows, audit


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Check WooCommerce prices against available real-price floors.")
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER)
    parser.add_argument("--woo-input", type=Path, default=DEFAULT_WOO_INPUT)
    parser.add_argument("--checked-output", type=Path, default=DEFAULT_CHECKED_OUTPUT)
    parser.add_argument("--strict-output", type=Path, default=DEFAULT_STRICT_OUTPUT)
    parser.add_argument("--audit-output", type=Path, default=DEFAULT_AUDIT_OUTPUT)
    parser.add_argument("--exceptions-output", type=Path, default=DEFAULT_EXCEPTIONS_OUTPUT)
    parser.add_argument("--supplier-matrix", type=Path, default=DEFAULT_SUPPLIER_MATRIX)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--usd-to-sar", type=float, default=DEFAULT_USD_TO_SAR)
    parser.add_argument("--unpublish-missing-reference", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    woo_columns, woo_rows = read_csv(args.woo_input)
    master_refs = load_master_refs(args.master, args.sheet)
    supplier_refs = load_supplier_refs(args.supplier_matrix, args.usd_to_sar)
    checked_rows, strict_rows, audit = audit_rows(
        woo_rows,
        master_refs,
        supplier_refs,
        args.usd_to_sar,
        unpublish_missing_reference=args.unpublish_missing_reference,
    )
    exceptions = [row for row in audit if row["Status"] != "PASS"]

    if rows_match(woo_rows, checked_rows, woo_columns):
        print(f"Skipped {args.checked_output}; no price-floor adjustments were needed.")
    else:
        write_csv(args.checked_output, woo_columns, checked_rows)
    write_csv(args.strict_output, woo_columns, strict_rows)
    write_csv(args.audit_output, AUDIT_COLUMNS, audit)
    write_csv(args.exceptions_output, AUDIT_COLUMNS, exceptions)

    counts: dict[str, int] = {}
    for row in audit:
        counts[row["Status"]] = counts.get(row["Status"], 0) + 1
    print(f"Checked {len(audit)} products.")
    for status in sorted(counts):
        print(f"{status}: {counts[status]}")
    print(f"Wrote {args.audit_output} and {args.exceptions_output}.")
    if not rows_match(woo_rows, checked_rows, woo_columns):
        print(f"Wrote checked import: {args.checked_output}.")
    print(f"Wrote strict import: {args.strict_output}.")


if __name__ == "__main__":
    main()
